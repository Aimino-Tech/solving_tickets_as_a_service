#!/usr/bin/env python3
"""Generates the AIM-4684 financial model (XLSX + CSV) for App Marketplace Arbitrage.

Base case: cumulative revenue >= EUR 100k by end of Week 4.
Sheets:
  1. Overview        — headline numbers, FX, EUR/USD conversion, target check
  2. Assumptions     — FX, fee rates, build cost, support cost, pricing
  3. Portfolio       — per-app: price, install base, rev-share, fee, revenue
  4. Weekly Ramp     — weekly new customers + cumulative revenue per app
  5. Sensitivities   — scenario table (unit counts x FX x mix)
  6. UnitEconomics   — per-app cost/margin summary
"""
from __future__ import annotations

import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).resolve().parent.parent / "docs" / "gtm" / "marketplace-arbitrage"

# ----------------------------------------------------------------------------
# Assumptions (single source of truth)
# ----------------------------------------------------------------------------
FX_EUR_USD = 1.08  # USD per EUR (base)
VAT_DE = 0.19  # German VAT on B2C digital sales (reverse-charge for B2B)

# Fee structure per store (researched 2026-08-06):
#   Shopify:    0% revenue share on first $1M lifetime (since 2025), 2.9% processing fee
#   HubSpot:    0% rev share, 0% processing if self-billed (Stripe ~2.9% if used)
#   AppExchange: 15% net revenue share (ISVforce) when sold via Checkout;
#                direct/custom installs pre-listing carry 0% share.
MARKET_FEES = {
    "Shopify": {"rev_share": 0.00, "processing": 0.029},
    "HubSpot": {"rev_share": 0.00, "processing": 0.029},
    "AppExchange": {"rev_share": 0.15, "processing": 0.00},
    "Slack": {"rev_share": 0.00, "processing": 0.029},
}

# Cost assumptions (USD)
BUILD_COST_PER_APP = 500  # AGI pipeline: compute + human review + QA
INFRA_COST_PER_APP = 50  # hosting, OAuth infra, monitoring per app per 4 weeks
SUPPORT_COST_WEEK = 800  # partial support FTE equivalent, per week of live apps

# Portfolio definition: price per year (USD), target install base by W4 (cumulative)
PORTFOLIO = [
    {"name": "App A — Sync-Fix Alternative", "store": "Shopify", "price_yr": 99,
     "cum_customers": 200, "setup_rate": 0.40, "setup_fee": 49},
    {"name": "App B — Price Alternative", "store": "Shopify", "price_yr": 149,
     "cum_customers": 150, "setup_rate": 0.30, "setup_fee": 79},
    {"name": "App C — HubSpot Gap", "store": "HubSpot", "price_yr": 490,
     "cum_customers": 60, "setup_rate": 0.50, "setup_fee": 149},
    {"name": "App D — Enterprise", "store": "AppExchange", "price_yr": 3500,
     "cum_customers": 8, "setup_rate": 1.00, "setup_fee": 750},
    {"name": "App E — Shopify Secondary", "store": "Shopify", "price_yr": 99,
     "cum_customers": 100, "setup_rate": 0.30, "setup_fee": 49},
]

# Weekly new-customer ramp per app (sums to cum_customers)
WEEKLY_RAMP = {
    "App A — Sync-Fix Alternative": [30, 50, 55, 65],
    "App B — Price Alternative": [20, 35, 45, 50],
    "App C — HubSpot Gap": [5, 12, 18, 25],
    "App D — Enterprise": [1, 2, 2, 3],
    "App E — Shopify Secondary": [10, 25, 30, 35],
}

# Sensitivity scenarios: (label, fx, unit_multiplier, enterprise_share_pct)
#   unit_multiplier scales all non-enterprise install bases
#   enterprise_share_pct = % of App D revenue sold via AppExchange Checkout
#   (rest sold direct pre-listing at 0% rev share)
SENSITIVITIES = [
    ("Base (5 Apps, 100% units)", 1.08, 1.00, 0.00),
    ("Pessimistisch (70% units)", 1.08, 0.70, 0.00),
    ("Optimistisch (120% units)", 1.08, 1.20, 0.00),
    ("FX 1.00 (EUR stark)", 1.00, 1.00, 0.00),
    ("FX 1.16 (EUR schwach)", 1.16, 1.00, 0.00),
    ("Enterprise via Checkout (15% RevShare)", 1.08, 1.00, 1.00),
    ("Worst Case (70% units, FX 1.00)", 1.00, 0.70, 0.00),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ACCENT_FILL = PatternFill("solid", fgColor="DDEBF7")
MONEY_FMT = '"$"#,##0'
EUR_FMT = '"EUR "#,##0'
PCT_FMT = "0.0%"


def money(v: float) -> float:
    return round(v, 2)


def app_revenue(app: dict, unit_mult: float = 1.0, fx: float = FX_EUR_USD) -> dict:
    """Revenue breakdown for one app under a scenario."""
    name = app["name"]
    store = app["store"]
    n = app["cum_customers"]
    if store == "AppExchange":
        n_units = n
    else:
        n_units = round(n * unit_mult)
    sub_rev = n_units * app["price_yr"]
    setup_rev = round(n_units * app["setup_rate"] * app["setup_fee"])
    gross = sub_rev + setup_rev
    fees = MARKET_FEES[store]
    rev_share = gross * fees["rev_share"]
    processing = sub_rev * fees["processing"]
    net = gross - rev_share - processing
    return {
        "name": name, "store": store, "n": n_units, "sub": sub_rev, "setup": setup_rev,
        "gross": gross, "rev_share": rev_share, "processing": processing,
        "net": net, "net_eur": net / fx,
    }


def total_for_scenario(fx: float, unit_mult: float, ent_checkout: float) -> dict:
    rows = []
    for app in PORTFOLIO:
        r = app_revenue(app, unit_mult, fx)
        # Apply enterprise checkout share only to AppExchange
        if app["store"] == "AppExchange":
            r["rev_share"] = r["gross"] * MARKET_FEES["AppExchange"]["rev_share"] * ent_checkout
            r["net"] = r["gross"] - r["rev_share"] - r["processing"]
            r["net_eur"] = r["net"] / fx
        rows.append(r)
    gross = sum(r["gross"] for r in rows)
    net = sum(r["net"] for r in rows)
    # COGS
    build = BUILD_COST_PER_APP * len(PORTFOLIO)
    infra = INFRA_COST_PER_APP * len(PORTFOLIO)
    live_weeks = {"App A — Sync-Fix Alternative": 4, "App B — Price Alternative": 4,
                  "App C — HubSpot Gap": 3, "App D — Enterprise": 3,
                  "App E — Shopify Secondary": 2}
    support = SUPPORT_COST_WEEK * sum(live_weeks.get(a["name"], 4) for a in PORTFOLIO)
    cogs = build + infra + support
    return {
        "rows": rows, "gross": gross, "net": net,
        "net_eur": net / fx, "cogs": cogs, "contribution": (net - cogs) / fx,
        "build": build, "infra": infra, "support": support,
    }


# ----------------------------------------------------------------------------
# Sheets
# ----------------------------------------------------------------------------
def style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def build_workbook() -> Workbook:
    wb = Workbook()

    # ---- Overview ----
    ws = wb.active
    ws.title = "Overview"
    base = total_for_scenario(FX_EUR_USD, 1.0, 0.0)
    ws.append(["App Marketplace Arbitrage — Finanzmodell (AIM-4684)"])
    ws.append(["Ziel: kumulierter Umsatz >= EUR 100.000 in 4 Wochen (W4)"])
    ws.append([])
    ws.append(["Metrik", "Wert", "Einheit"])
    style_header(ws, 4, 3)
    ws.append(["Wechselkurs EUR/USD", FX_EUR_USD, "USD/EUR"])
    ws.append(["Bruttoumsatz (USD)", base["gross"], "USD"])
    ws.append(["Marktgebühren (RevShare+Processing)", round(sum(r["rev_share"] + r["processing"] for r in base["rows"]), 2), "USD"])
    ws.append(["Netto-Umsatz (USD)", base["net"], "USD"])
    ws.append(["Netto-Umsatz (EUR)", base["net_eur"], "EUR"])
    ws.append(["COGS (Build+Infra+Support)", base["cogs"], "USD"])
    ws.append(["Contribution (EUR)", base["contribution"], "EUR"])
    ws.append(["Ziel erreicht (>= EUR 100k Netto)?", "JA" if base["net_eur"] >= 100_000 else "NEIN", ""])
    ws.append([])
    ws.append(["Hinweis: Netto-Umsatz = Brutto abzgl. RevShare und Processing-Gebühren."])
    ws.append(["EUR-Beträge = USD / FX. MwSt (DE 19%) ist bei B2B-Reverse-Charge nicht enthalten."])
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12

    # ---- Assumptions ----
    ws = wb.create_sheet("Assumptions")
    ws.append(["Annahme", "Wert", "Quelle / Hinweis"])
    style_header(ws, 1, 3)
    ws.append(["FX EUR/USD", FX_EUR_USD, "Sensitivität: 1.00–1.16"])
    ws.append(["MwSt DE (B2C)", VAT_DE, "B2B-Reverse-Charge: 0%"])
    ws.append(["Shopify RevShare", 0.00, "0% auf erste $1M lifetime (seit 2025)"])
    ws.append(["Shopify Processing", 0.029, "2.9% auf Billing-Volumen"])
    ws.append(["HubSpot RevShare", 0.00, "0% (keine RevShare, kein Listing-Fee)"])
    ws.append(["HubSpot Processing", 0.029, "nur wenn Stripe genutzt"])
    ws.append(["AppExchange RevShare", 0.15, "15% ISVforce bei Checkout; 0% bei Direct/Pre-Listing"])
    ws.append(["AppExchange Security Review", 999, "USD pro Attempt (je 2 Attempts üblich)"])
    ws.append(["Build-Kosten/App", BUILD_COST_PER_APP, "USD (AGI-Pipeline 48h + Review)"])
    ws.append(["Infra-Kosten/App", INFRA_COST_PER_APP, "USD über 4 Wochen"])
    ws.append(["Support-Kosten/Woche", SUPPORT_COST_WEEK, "USD (partielle FTE)"])
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 52

    # ---- Portfolio ----
    ws = wb.create_sheet("Portfolio")
    ws.append(["App", "Store", "Kunden (W4)", "Preis/Jahr", "Setup-Quote", "Setup-Fee",
               "Subscription (USD)", "Setup (USD)", "Brutto (USD)", "RevShare (USD)",
               "Processing (USD)", "Netto (USD)", "Netto (EUR)"])
    style_header(ws, 1, 13)
    base_rows = total_for_scenario(FX_EUR_USD, 1.0, 0.0)["rows"]
    for r in base_rows:
        ws.append([r["name"], r["store"], r["n"], r["sub"] // max(r["n"], 1),
                   PORTFOLIO[[p["name"] for p in PORTFOLIO].index(r["name"])]["setup_rate"],
                   PORTFOLIO[[p["name"] for p in PORTFOLIO].index(r["name"])]["setup_fee"],
                   r["sub"], r["setup"], r["gross"], r["rev_share"], r["processing"],
                   r["net"], r["net_eur"]])
    ws.append(["SUMME", "", "", "", "", "",
               sum(r["sub"] for r in base_rows), sum(r["setup"] for r in base_rows),
               sum(r["gross"] for r in base_rows), sum(r["rev_share"] for r in base_rows),
               sum(r["processing"] for r in base_rows), sum(r["net"] for r in base_rows),
               round(sum(r["net_eur"] for r in base_rows), 2)])
    for cell in ws[ws.max_row]:
        cell.fill = ACCENT_FILL
        cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 30
    for c in range(2, 14):
        ws.column_dimensions[get_column_letter(c)].width = 15

    # ---- Weekly Ramp ----
    ws = wb.create_sheet("Weekly Ramp")
    ws.append(["App", "W1 Neu", "W2 Neu", "W3 Neu", "W4 Neu", "Kumuliert W4",
               "Ø-Preis/Jahr", "Kum. Netto-Umsatz (EUR)"])
    style_header(ws, 1, 8)
    grand_net_eur = 0.0
    grand_cum = 0
    for app in PORTFOLIO:
        ramp = WEEKLY_RAMP[app["name"]]
        cum = sum(ramp)
        # net eur for this app at full unit count
        r = app_revenue(app, 1.0, FX_EUR_USD)
        net_eur = r["net_eur"]
        grand_net_eur += net_eur
        grand_cum += cum
        ws.append([app["name"], *ramp, cum, app["price_yr"], round(net_eur, 2)])
    ws.append(["SUMME", "", "", "", "", grand_cum, "", round(grand_net_eur, 2)])
    for cell in ws[ws.max_row]:
        cell.fill = ACCENT_FILL
        cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 30
    for c in range(2, 9):
        ws.column_dimensions[get_column_letter(c)].width = 16

    # ---- Sensitivities ----
    ws = wb.create_sheet("Sensitivities")
    ws.append(["Szenario", "FX", "Unit-Multiplikator", "Enterprise-Checkout-Anteil",
               "Brutto (USD)", "Netto (USD)", "Netto (EUR)", "COGS (USD)",
               "Contribution (EUR)", ">= 100k?"])
    style_header(ws, 1, 10)
    for label, fx, mult, ent in SENSITIVITIES:
        t = total_for_scenario(fx, mult, ent)
        hit = "JA" if t["net_eur"] >= 100_000 else "NEIN"
        ws.append([label, fx, mult, ent, t["gross"], t["net"], round(t["net_eur"], 2),
                   t["cogs"], round(t["contribution"], 2), hit])
    ws.column_dimensions["A"].width = 42
    for c in range(2, 11):
        ws.column_dimensions[get_column_letter(c)].width = 18

    # ---- Unit Economics ----
    ws = wb.create_sheet("UnitEconomics")
    ws.append(["App", "Store", "Kunden", "Ø-Preis/Jahr", "Brutto/Kunde",
               "RevShare/Kunde", "Processing/Kunde", "Netto/Kunde", "COGS-Quote",
               "Netto-Marge (Kunde)"])
    style_header(ws, 1, 10)
    for r in base_rows:
        price_idx = [p["name"] for p in PORTFOLIO].index(r["name"])
        price = PORTFOLIO[price_idx]["price_yr"]
        setup = PORTFOLIO[price_idx]["setup_rate"] * PORTFOLIO[price_idx]["setup_fee"]
        gross_pc = price + setup
        rev_pc = r["rev_share"] / max(r["n"], 1)
        proc_pc = r["processing"] / max(r["n"], 1)
        net_pc = gross_pc - rev_pc - proc_pc
        ws.append([r["name"], r["store"], r["n"], price, round(gross_pc, 2),
                   round(rev_pc, 2), round(proc_pc, 2), round(net_pc, 2),
                   round((BUILD_COST_PER_APP / max(r["n"], 1)), 2),
                   round(net_pc / gross_pc, 3)])
    ws.column_dimensions["A"].width = 30
    for c in range(2, 11):
        ws.column_dimensions[get_column_letter(c)].width = 16

    return wb


def write_csv(out: Path, scenario_label: str = "Base") -> None:
    t = total_for_scenario(FX_EUR_USD, 1.0, 0.0)
    with open(out, "w", newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["App Marketplace Arbitrage — Finanzmodell", "Szenario:", scenario_label])
        w.writerow([])
        w.writerow(["App", "Store", "Kunden", "Sub (USD)", "Setup (USD)", "Brutto (USD)",
                    "RevShare (USD)", "Processing (USD)", "Netto (USD)", "Netto (EUR)"])
        for r in t["rows"]:
            w.writerow([r["name"], r["store"], r["n"], r["sub"], r["setup"], r["gross"],
                        r["rev_share"], r["processing"], r["net"], round(r["net_eur"], 2)])
        w.writerow(["SUMME", "", "", sum(r["sub"] for r in t["rows"]),
                    sum(r["setup"] for r in t["rows"]), t["gross"],
                    sum(r["rev_share"] for r in t["rows"]),
                    sum(r["processing"] for r in t["rows"]), t["net"], round(t["net_eur"], 2)])
        w.writerow([])
        w.writerow(["COGS (USD)", t["cogs"], "Contribution (EUR)", round(t["contribution"], 2)])
        w.writerow(["FX EUR/USD", FX_EUR_USD, "Netto EUR", round(t["net_eur"], 2)])


def main() -> None:
    wb = build_workbook()
    os.makedirs(OUT_DIR, exist_ok=True)
    xlsx = OUT_DIR / "financial-model.xlsx"
    csvp = OUT_DIR / "financial-model.csv"
    wb.save(xlsx)
    write_csv(csvp)
    base = total_for_scenario(FX_EUR_USD, 1.0, 0.0)
    print(f"XLSX: {xlsx}")
    print(f"CSV:  {csvp}")
    print(f"Base gross USD:  {base['gross']:,.0f}")
    print(f"Base net USD:    {base['net']:,.0f}")
    print(f"Base net EUR:    {base['net_eur']:,.0f}")
    print(f"Target >= 100k EUR: {'JA' if base['net_eur'] >= 100_000 else 'NEIN'}")
    for label, fx, mult, ent in SENSITIVITIES:
        t = total_for_scenario(fx, mult, ent)
        print(f"  [SEN] {label:45s} net EUR {t['net_eur']:10,.0f}  {'OK' if t['net_eur'] >= 100_000 else 'UNTER'}")


if __name__ == "__main__":
    main()
